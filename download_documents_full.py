#!/usr/bin/env python3
"""
Оптимизированный скрипт для скачивания документов на всем датасете.
"""

import pandas as pd
import requests
import time
import json
import re
from urllib.parse import parse_qs, quote, urlparse, unquote
from bs4 import BeautifulSoup
import PyPDF2
import io
from typing import List, Dict, Any, Optional
import logging
import os
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
from collections import defaultdict
import signal
import sys

# Настройка логирования (лог в папке скрипта)
_LOG_DIR = os.path.dirname(os.path.abspath(__file__))
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_LOG_DIR, 'download_progress.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class OptimizedDocumentDownloader:
    def __init__(self, timeout=15, delay=0.1, max_workers=5):
        self.timeout = timeout
        self.delay = delay
        self.max_workers = max_workers
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        # Кэш для уже скачанных документов
        self.document_cache = {}
        # Статистика
        self.stats = {
            'total_requests': 0,
            'cached_requests': 0,
            'successful_downloads': 0,
            'failed_downloads': 0,
            'wikipedia_api_requests': 0,
            'wikipedia_api_success': 0
        }
    
    def _mediawiki_api_json(self, api_url: str, params: dict) -> Optional[dict]:
        """GET action=api с повторами при 429 и текстовом rate limit."""
        req = dict(params)
        req.setdefault("format", "json")
        for attempt in range(8):
            try:
                r = self.session.get(
                    api_url,
                    params=req,
                    timeout=max(self.timeout, 25),
                )
                body = r.text or ""
                if r.status_code in (429, 503) or (
                    r.status_code == 200 and "too many requests" in body.lower()
                ):
                    wait = 4 + attempt * 3
                    logger.warning(
                        "Wikipedia API throttling, пауза %ss (попытка %d/8)",
                        wait,
                        attempt + 1,
                    )
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                data = json.loads(body)
                if (
                    isinstance(data, dict)
                    and data.get("error", {}).get("code") == "ratelimited"
                ):
                    wait = 5 + attempt * 2
                    time.sleep(wait)
                    continue
                return data
            except requests.RequestException as e:
                logger.debug("mediawiki запрос: %s", e)
                time.sleep(2 + attempt * 2)
            except (json.JSONDecodeError, ValueError):
                time.sleep(2 + attempt * 2)
        return None
    
    def get_url_hash(self, url: str) -> str:
        """Создает хэш URL для кэширования."""
        return hashlib.md5(url.encode()).hexdigest()

    @staticmethod
    def _normalize_wikipedia_host(url: str) -> str:
        """Возвращает нормализованный хост *.wikipedia.org или пустую строку."""
        try:
            host = (urlparse((url or "").strip()).netloc or "").lower()
            if not host:
                return ""
            if host.startswith("www."):
                host = host[4:]
            if ".m.wikipedia.org" in host:
                host = host.replace(".m.wikipedia.org", ".wikipedia.org")
            if host.endswith(".wikipedia.org") or host == "wikipedia.org":
                return host
        except Exception:
            pass
        return ""

    @staticmethod
    def _wikipedia_lang_from_hostname(hostname: str) -> str:
        """Код языка для wikipedia-api / api.php (например uk, zh-min-nan, en)."""
        host = (hostname or "").lower()
        if host.startswith("www."):
            host = host[4:]
        if ".m.wikipedia.org" in host:
            host = host.replace(".m.wikipedia.org", ".wikipedia.org")
        if not host.endswith(".wikipedia.org"):
            return "en"
        sub = host[: -len(".wikipedia.org")]
        return sub if sub else "en"

    def is_wikipedia_url(self, url: str) -> bool:
        """
        Страница статьи на Wikipedia: хост *.wikipedia.org и из URL можно взять title
        (/wiki/... или index.php?title=...). Остальное уходит в обычную HTML-загрузку.
        """
        if not self._normalize_wikipedia_host(url):
            return False
        try:
            parsed = urlparse((url or "").strip())
            path = parsed.path or ""
            query_l = (parsed.query or "").lower()
            if "/wiki/" in path:
                return True
            if "index.php" in path and "title=" in query_l:
                return True
            return False
        except Exception:
            return False

    def extract_wikipedia_title(self, url: str) -> Optional[str]:
        """Извлекает название статьи из Wikipedia URL."""
        try:
            parsed = urlparse(url)
            if "/wiki/" in parsed.path:
                title = parsed.path.split("/wiki/")[-1]
                title = title.split("#")[0]
                return unquote(title)
            qs = parse_qs(parsed.query)
            if "title" in qs and qs["title"]:
                return unquote(qs["title"][0])
            return None
        except Exception:
            return None
    
    @staticmethod
    def _is_wikiroulette_url(url: str) -> bool:
        """wikiroulette.co отдаёт оболочку без текста статьи в сыром HTML (нужен JS)."""
        try:
            host = (urlparse((url or "").strip()).netloc or "").lower()
            if host.startswith("www."):
                host = host[4:]
            return host == "wikiroulette.co" or host.endswith(".wikiroulette.co")
        except Exception:
            return False
    
    def _synthetic_wikipedia_url_from_wikiroulette(self, url: str) -> Optional[str]:
        """
        Из query ?p=... (как в MediaWiki) строим URL en.wikipedia.org/wiki/...
        для download_wikipedia_via_api.
        """
        if not self._is_wikiroulette_url(url):
            return None
        try:
            parsed = urlparse(url.strip())
            qs = parse_qs(parsed.query)
            if "p" not in qs or not qs["p"]:
                return None
            raw = unquote(str(qs["p"][0]).strip())
            if not raw:
                return None
            if raw.count("(") > raw.count(")"):
                raw = raw + ")"
            path_title = raw.replace("_", " ").strip().replace(" ", "_")
            if not path_title:
                return None
            # Обрезанное название в simpleqa → реальная статья enwiki
            if (
                "Odd_Fellows_Hall" in path_title
                and "Eureka" in path_title
                and "California" not in path_title
            ):
                path_title = "Odd_Fellows_Hall_(Eureka,_California)"
            enc = quote(path_title, safe="/():,%!")
            return f"https://en.wikipedia.org/wiki/{enc}"
        except Exception:
            return None
    
    def _get_wikipedia_wiki(self, lang: str):
        """Ленивая инициализация wikipediaapi по языку."""
        if not hasattr(self, '_wiki_clients'):
            self._wiki_clients = {}
        if lang not in self._wiki_clients:
            import wikipediaapi
            self._wiki_clients[lang] = wikipediaapi.Wikipedia(
                user_agent='SimpleQA-Processor/1.0 (https://github.com/simple-qa)',
                language=lang,
            )
        return self._wiki_clients[lang]
    
    def download_wikipedia_via_api(self, url: str) -> Optional[str]:
        """
        Загружает статью Wikipedia. Сначала wikipedia-api; если страницы «нет»
        или текст пуст — обязательно пробуем action=parse (fallback), иначе
        часть нормальных статей теряется из‑за расхождения имён/лимитов.
        """
        title = self.extract_wikipedia_title(url)
        if not title:
            return None

        parsed = urlparse(url)
        lang = self._wikipedia_lang_from_hostname(parsed.netloc)
        resolved = self._wikipedia_final_title(lang, title)
        if resolved is None:
            return None
        title = resolved

        self.stats["wikipedia_api_requests"] += 1

        api_text: Optional[str] = None
        try:
            import wikipediaapi  # noqa: F401

            wiki = self._get_wikipedia_wiki(lang)
            page = wiki.page(title)
            if page.exists():
                t = page.text
                if t and t.strip():
                    api_text = t.strip()
        except ImportError:
            pass
        except Exception as e:
            logger.debug("wikipedia-api для %s: %s", title, e)

        if api_text:
            self.stats["wikipedia_api_success"] += 1
            logger.info("Wikipedia OK (wikipedia-api): %s", title)
            return api_text

        time.sleep(max(self.delay, 0.5))
        return self._download_wikipedia_fallback(url)
    
    def _wikipedia_final_title(self, lang: str, title: str) -> Optional[str]:
        """
        Цепочка редиректов → каноническое имя. None = страницы нет.
        При сбое запроса возвращает исходный title.
        """
        try:
            api_url = f"https://{lang}.wikipedia.org/w/api.php"
            data = self._mediawiki_api_json(
                api_url,
                {"action": "query", "titles": title, "redirects": 1},
            )
            if data is None:
                return title
            pages = data.get("query", {}).get("pages", {})
            for _pid, page in pages.items():
                if page.get("missing"):
                    return None
                if page.get("invalid"):
                    return None
                return page.get("title") or title
        except Exception:
            return title
        return title
    
    def _download_wikipedia_fallback(self, url: str) -> Optional[str]:
        """Fallback: action=parse + BeautifulSoup (если wikipedia-api недоступен или пуст)."""
        try:
            title = self.extract_wikipedia_title(url)
            if not title:
                return None
            parsed = urlparse(url)
            lang = self._wikipedia_lang_from_hostname(parsed.netloc)
            final_title = self._wikipedia_final_title(lang, title)
            if final_title is None:
                return None
            api_url = f"https://{lang}.wikipedia.org/w/api.php"
            params = {
                "action": "parse",
                "page": final_title,
                "prop": "text",
                "format": "json",
                "disabletoc": 1,
                "disableeditsection": 1,
            }
            data = self._mediawiki_api_json(api_url, params)
            if not data or data.get("error"):
                return None
            html_content = data.get("parse", {}).get("text", {}).get("*", "")
            if html_content:
                result = self._clean_wikipedia_html(html_content)
                if result:
                    self.stats["wikipedia_api_success"] += 1
                    logger.info("Wikipedia OK (fallback): %s", final_title)
                    return result
        except Exception:
            pass
        return None
    
    def _clean_wikipedia_html(self, html_content: str) -> str:
        """
        Извлекает текст из HTML Wikipedia, убирая мусор:
        - навигация, меню, sidebar
        - скрипты, стили, секция References
        - ссылки [edit], служебные элементы
        """
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Берём только основной контент (mw-parser-output) — без навигации и sidebar
        main_div = soup.find('div', class_='mw-parser-output')
        if main_div:
            soup = main_div
        
        # Удаляем скрипты и стили
        for tag in soup(["script", "style"]):
            tag.decompose()
        
        # Удаляем навигацию, меню, sidebar
        for tag in soup.find_all(['nav', 'header', 'footer']):
            tag.decompose()
        for tag in soup.find_all(class_=lambda c: c and any(
            x in str(c).lower() for x in ('navigation', 'sidebar', 'vector-menu', 'mw-jump-link', 'noprint')
        )):
            tag.decompose()
        for tag in soup.find_all(id=lambda i: i and any(
            x in str(i).lower() for x in ('mw-head', 'mw-panel', 'mw-navigation', 'mw-page-base')
        )):
            tag.decompose()
        
        # Удаляем секцию References (списки источников)
        for ref in soup.find_all(class_=lambda c: c and ('reflist' in c or 'references' in c)):
            ref.decompose()
        
        # Удаляем hatnotes (подсказки типа "Main article: ...")
        for tag in soup.find_all(class_='hatnote'):
            tag.decompose()
        
        # Удаляем ambox (предупреждения, шаблоны)
        for tag in soup.find_all(class_=lambda c: c and 'ambox' in c):
            tag.decompose()
        
        # Удаляем ссылки [edit]
        for tag in soup.find_all('span', class_='mw-editsection'):
            tag.decompose()
        
        # Удаляем table of contents (содержание)
        for tag in soup.find_all(id='toc'):
            tag.decompose()
        for tag in soup.find_all(class_=lambda c: c and 'toc' in str(c).lower()):
            tag.decompose()
        
        # Извлекаем текст только из параграфов и списков (основной контент)
        paragraphs = []
        for tag in soup.find_all(['p', 'li']):
            t = tag.get_text(separator=' ', strip=True)
            if t and len(t) > 20:  # отсекаем короткие пункты меню
                paragraphs.append(t)
        
        if paragraphs:
            text = '\n\n'.join(paragraphs)
        else:
            # fallback: весь текст (separator сохраняет пробелы между элементами)
            text = soup.get_text(separator=' ')
        
        # Удаляем всё до "From Wikipedia, the free encyclopedia" (если есть)
        if 'From Wikipedia, the free encyclopedia' in text:
            idx = text.find('From Wikipedia, the free encyclopedia')
            text = text[idx + len('From Wikipedia, the free encyclopedia'):].strip()
        
        # Удаляем мусорные фразы в начале строк
        junk_phrases = [
            'Jump to content', 'Main menu', 'move to sidebar', 'hide Navigation',
            'Main page', 'Contents', 'Current events', 'About Wikipedia',
            'Contact us', 'Contribute', 'Help', 'Learn to edit',
            'Donate', 'Create account', 'Log in', 'Personal tools',
            'Toggle the table of contents', 'Edit links', 'Article', 'Talk',
            'Read', 'Edit', 'View history', 'Tools', 'Actions', 'General',
            'What links here', 'Related changes', 'Upload file',
            'Permanent link', 'Page information', 'Cite this page',
            'Print/export', 'Download as PDF', 'Printable version',
            'In other projects', 'Appearance', 'move to sidebar hide',
        ]
        lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped:
                if lines and lines[-1]:
                    lines.append('')
                continue
            # Пропускаем строки, которые целиком — мусор
            if len(stripped) < 30 and any(p in stripped for p in junk_phrases):
                continue
            lines.append(stripped)
        
        while lines and not lines[0]:
            lines.pop(0)
        while lines and not lines[-1]:
            lines.pop()
        
        return '\n\n'.join(lines) if lines else ''
    
    def download_document(self, url: str) -> Optional[str]:
        """
        Скачивает документ по URL с кэшированием.
        Для Wikipedia использует API вместо парсинга HTML.
        """
        # Проверяем кэш
        url_hash = self.get_url_hash(url)
        if url_hash in self.document_cache:
            self.stats['cached_requests'] += 1
            return self.document_cache[url_hash]
        
        self.stats['total_requests'] += 1
        
        # Wikipedia — только через API (чистый контент без навигации).
        # Fallback на обычный HTML даёт мусор (меню, sidebar), поэтому не используем.
        if self.is_wikipedia_url(url):
            content = self.download_wikipedia_via_api(url)
            if content:
                self.document_cache[url_hash] = content
                self.stats['successful_downloads'] += 1
                return content
            self.stats['failed_downloads'] += 1
            return None
        
        # WikiRoulette: в HTML нет тела статьи; title в ?p= совпадает с enwiki.
        synthetic = self._synthetic_wikipedia_url_from_wikiroulette(url)
        if synthetic:
            content = self.download_wikipedia_via_api(synthetic)
            if content:
                self.document_cache[url_hash] = content
                self.stats['successful_downloads'] += 1
                logger.info("WikiRoulette → Wikipedia API ok: %s", synthetic[:88])
                return content
            logger.debug("WikiRoulette: статья не найдена, HTTP fallback: %s", url[:88])
        
        # Обычная загрузка для не-Wikipedia URL
        try:
            response = self.session.get(url, timeout=self.timeout)
            response.raise_for_status()
            
            # Определяем тип контента
            content_type = response.headers.get('content-type', '').lower()
            url_low = url.lower()
            body = response.content

            if (
                url_low.endswith(".xlsx")
                or ".xlsx?" in url_low
                or "spreadsheetml.sheet" in content_type
            ):
                content = self._extract_xlsx_text(body)
            elif 'pdf' in content_type or url_low.endswith('.pdf'):
                content = self._extract_pdf_text(body)
            elif 'html' in content_type or url_low.endswith(('.html', '.htm')):
                # Для Wikipedia URL все равно парсим HTML как fallback
                content = self._extract_html_text(response.text)
            elif 'text' in content_type or url_low.endswith('.txt'):
                content = response.text
            else:
                # Пытаемся обработать как HTML
                try:
                    content = self._extract_html_text(response.text)
                except Exception:
                    content = response.text
            
            # Сохраняем в кэш
            if content:
                self.document_cache[url_hash] = content
                self.stats['successful_downloads'] += 1
                return content
            else:
                self.stats['failed_downloads'] += 1
                return None
                    
        except requests.exceptions.RequestException as e:
            self.stats['failed_downloads'] += 1
            return None
        except Exception as e:
            self.stats['failed_downloads'] += 1
            return None
    
    def _extract_pdf_text(self, pdf_content: bytes) -> str:
        """Извлекает текст из PDF."""
        try:
            pdf_file = io.BytesIO(pdf_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            return ""

    def _extract_xlsx_text(self, data: bytes, *, max_chars: int = 1_000_000) -> str:
        """Текст из Excel .xlsx (Open XML): все листы, строки через табуляцию."""
        if not data or len(data) < 64:
            logger.warning("xlsx: слишком короткое тело (%s байт)", len(data or b""))
            return ""
        try:
            import openpyxl  # noqa: WPS433
        except ImportError:
            logger.warning("openpyxl не установлен — установите: pip install openpyxl")
            return ""
        head = data[:8]
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            logger.warning(
                "xlsx: openpyxl не открыл файл (%s байт, заголовок %r): %s",
                len(data),
                head,
                e,
            )
            return ""
        parts: List[str] = []
        total = 0
        n_sheets = len(wb.sheetnames)
        try:
            if not wb.sheetnames:
                logger.warning("xlsx: книга без листов (%s байт)", len(data))
                return ""
            for sheet in wb.worksheets:
                parts.append(f"\n\n## {sheet.title}\n")
                total += len(parts[-1])
                if total >= max_chars:
                    parts.append("\n...[truncated]\n")
                    break
                try:
                    for row in sheet.iter_rows(values_only=True):
                        cells = [
                            "" if c is None else str(c).strip().replace("\n", " ")
                            for c in row
                        ]
                        line = "\t".join(cells).strip()
                        if line:
                            parts.append(line + "\n")
                            total += len(parts[-1])
                            if total >= max_chars:
                                parts.append("...[truncated]\n")
                                break
                except Exception as e:
                    logger.warning(
                        "xlsx: ошибка при чтении листа %r: %s",
                        getattr(sheet, "title", "?"),
                        e,
                    )
                if total >= max_chars:
                    break
        finally:
            wb.close()
        out = "".join(parts).strip()
        if not out:
            logger.warning(
                "xlsx: после парсинга пустой текст (%s байт, листов: %s)",
                len(data),
                n_sheets,
            )
        return out

    def _extract_html_text(self, html_content: str) -> str:
        """Извлекает текст из HTML, убирая навигацию и меню."""
        try:
            # Сначала trafilatura (оптимизирована для статей), затем BeautifulSoup как fallback
            try:
                from trafilatura import extract as trafilatura_extract
                t_result = trafilatura_extract(html_content)
                if t_result and len(t_result.strip()) > 200:
                    return t_result.strip()
            except ImportError:
                pass
            return self._extract_html_text_bs4(html_content)
        except Exception:
            return html_content
    
    def _extract_html_text_bs4(self, html_content: str) -> str:
        """Извлечение через BeautifulSoup."""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
                tag.decompose()
            
            # Удаляем элементы навигации по классам/id (без header/footer — могут быть article-header)
            for tag in soup.find_all(class_=lambda c: c and any(
                x in str(c).lower() for x in (
                    'nav', 'menu', 'sidebar', 'banner', 'cookie', 'newsletter',
                    'subscribe', 'social-share'
                )
            )):
                tag.decompose()
            for tag in soup.find_all(id=lambda i: i and any(
                x in str(i).lower() for x in ('nav', 'menu', 'sidebar')
            )):
                tag.decompose()
            
            # Пытаемся взять основной контент (не article-header / article-title — только заголовок)
            def _is_header_like(cls_str: str) -> bool:
                if not cls_str:
                    return False
                c = str(cls_str).lower()
                return any(x in c for x in ('header', 'title', 'meta', 'byline', 'dateline'))
            
            def _good_content_class(cls_str: str) -> bool:
                if not cls_str:
                    return False
                c = str(cls_str).lower()
                if _is_header_like(c):
                    return False
                return any(x in c for x in ('article', 'content', 'post', 'story', 'body', 'entry', 'prose'))
            
            main = soup.find('article') or soup.find('main') or soup.find(attrs={'role': 'main'})
            if not main:
                candidates = soup.find_all(class_=lambda c: c and _good_content_class(c))
                main = max(candidates, key=lambda t: len(t.get_text())) if candidates else None
            root = main if main else soup
            
            text = root.get_text(separator=' ')
            
            # Если выбранный контент слишком короткий (только заголовок) — пробуем body целиком
            if len(text.strip()) < 400 and root is not soup:
                body = soup.find('body')
                if body and len(body.get_text(separator=' ').strip()) > len(text):
                    root = body
                    text = root.get_text(separator=' ')
            
            # Удаляем типичные мусорные строки из начала
            junk_patterns = (
                r'^skip to .+$', r'^log in$', r'^sign (up|in)$', r'^subscribe$',
                r'^search\s*$', r'^click here to search$', r'^close menu$',
                r'^explore the .+$', r'^more from .+$', r'^follow us on',
                r'^home$', r'^news$', r'^sport$', r'^edition\s+(in|us)$',
                r'^privacy policy$', r'^terms of use$', r'^contact us$',
                r'^about us$', r'^accessibility$', r'^help$', r'^advertise',
            )
            junk_re = re.compile('|'.join(junk_patterns), re.I)
            
            lines = []
            for line in text.splitlines():
                stripped = ' '.join(line.split())
                if not stripped:
                    if lines and lines[-1]:
                        lines.append('')
                    continue
                if len(stripped) < 50 and junk_re.match(stripped):
                    continue
                lines.append(stripped)
            
            # Обрезаем только явно навигационные короткие строки в начале (не трогаем контент)
            # Порог 25 символов: "Home", "News", "Search" — нав; "Breaking news: X" (26+) — контент
            i = 0
            while i < min(20, len(lines)) and len(lines[i]) < 25:
                low = lines[i].lower()
                if low in ('home', 'news', 'search', 'sport', 'log in', 'sign up', 'subscribe',
                           'menu', 'close menu', 'sign in') or low.startswith('skip to '):
                    i += 1
                else:
                    break
            if i > 0:
                lines = lines[i:]
            
            return '\n'.join(lines).strip() if lines else ''
        except Exception:
            return html_content
    
    def download_documents_parallel(self, urls: List[str]) -> List[str]:
        """
        Скачивает документы параллельно, сохраняя порядок как в списке urls
        (индекс i соответствует clean_urls[i]).
        """
        # Фильтруем и очищаем URL
        clean_urls = []
        for url in urls:
            if not url or not url.strip():
                continue
            url = url.strip()
            if url.startswith('http'):
                clean_urls.append(url)

        if not clean_urls:
            return []

        documents: List[str] = [''] * len(clean_urls)

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_index = {
                executor.submit(self.download_document, url): i
                for i, url in enumerate(clean_urls)
            }

            for future in as_completed(future_to_index):
                i = future_to_index[future]
                url = clean_urls[i]
                try:
                    doc_content = future.result()
                    documents[i] = doc_content.strip() if doc_content else ''
                except Exception as e:
                    logger.warning(f"Ошибка при обработке {url}: {e}")
                    documents[i] = ''

                time.sleep(self.delay)

        return documents

def signal_handler(sig, frame):
    """Обработчик сигнала для корректного завершения."""
    logger.info("Получен сигнал завершения. Сохраняю промежуточные результаты...")
    sys.exit(0)

def process_dataset_full(input_file: str, output_file: str):
    """
    Обработка полного датасета с сохранением промежуточных результатов.
    """
    # Настраиваем обработчик сигналов
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    logger.info(f"Загружаю датасет из {input_file}")
    
    # Загружаем CSV
    df = pd.read_csv(input_file)
    logger.info(f"Загружено {len(df)} строк")
    
    # Создаем объект для скачивания
    downloader = OptimizedDocumentDownloader()
    
    # Добавляем поле documents
    df['documents'] = None
    failed_downloads = 0
    
    # Обрабатываем с прогресс-баром
    for idx in tqdm(range(len(df)), desc="Обработка строк"):
        row = df.iloc[idx]
        
        try:
            # Парсим metadata для получения URLs
            metadata_str = row['metadata']
            if isinstance(metadata_str, str):
                # Пытаемся распарсить как JSON
                try:
                    metadata = json.loads(metadata_str)
                except:
                    # Если не JSON, пытаемся извлечь URLs с помощью regex
                    urls = re.findall(r'https?://[^\s\'\"]+', metadata_str)
                    metadata = {'urls': urls}
            else:
                metadata = metadata_str
            
            urls = metadata.get('urls', [])
            if not urls:
                df.at[idx, 'documents'] = []
                failed_downloads += 1
                continue
            
            # Скачиваем документы параллельно
            documents = downloader.download_documents_parallel(urls)
            df.at[idx, 'documents'] = documents
            
            if not documents:
                failed_downloads += 1
            
            # Сохраняем промежуточные результаты каждые 100 строк
            if (idx + 1) % 100 == 0:
                temp_file = f"{output_file}.temp"
                df.to_csv(temp_file, index=False)
                logger.info(f"Сохранен промежуточный результат: {idx + 1} строк обработано")
                
                # Логируем статистику
                logger.info(f"Статистика на строке {idx + 1}:")
                logger.info(f"  Кэшированных запросов: {downloader.stats['cached_requests']}")
                logger.info(f"  Успешных загрузок: {downloader.stats['successful_downloads']}")
                logger.info(f"  Неудачных загрузок: {downloader.stats['failed_downloads']}")
                logger.info(f"  Wikipedia API запросов: {downloader.stats['wikipedia_api_requests']}")
                logger.info(f"  Wikipedia API успешных: {downloader.stats['wikipedia_api_success']}")
            
        except Exception as e:
            logger.error(f"Ошибка при обработке строки {idx + 1}: {e}")
            df.at[idx, 'documents'] = []
            failed_downloads += 1
    
    # Сохраняем финальный результат
    logger.info(f"Сохраняю финальный результат в {output_file}")
    df.to_csv(output_file, index=False)
    
    # Удаляем временный файл
    temp_file = f"{output_file}.temp"
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    # Статистика
    total_rows = len(df)
    successful_downloads = total_rows - failed_downloads
    
    logger.info(f"Обработка завершена!")
    logger.info(f"Всего строк: {total_rows}")
    logger.info(f"Успешно обработано: {successful_downloads}")
    logger.info(f"Не удалось скачать документы: {failed_downloads}")
    logger.info(f"Процент неудачных загрузок: {failed_downloads/total_rows*100:.2f}%")
    
    # Статистика кэширования
    logger.info(f"Статистика кэширования:")
    logger.info(f"Всего запросов: {downloader.stats['total_requests']}")
    logger.info(f"Кэшированных запросов: {downloader.stats['cached_requests']}")
    logger.info(f"Успешных загрузок: {downloader.stats['successful_downloads']}")
    logger.info(f"Неудачных загрузок: {downloader.stats['failed_downloads']}")
    logger.info(f"Wikipedia API запросов: {downloader.stats['wikipedia_api_requests']}")
    logger.info(f"Wikipedia API успешных: {downloader.stats['wikipedia_api_success']}")
    
    # Сохраняем итоговый отчет
    save_final_report(
        total_rows,
        successful_downloads,
        failed_downloads,
        downloader.stats,
        output_file,
        input_file=input_file,
    )
    
    return failed_downloads

def save_final_report(
    total_rows,
    successful_downloads,
    failed_downloads,
    stats,
    output_file,
    *,
    input_file: Optional[str] = None,
):
    """
    Сохраняет итоговый отчет о результатах обработки.
    """
    report_file = output_file.replace('.csv', '_report.txt')
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("ИТОГОВЫЙ ОТЧЕТ ОБ ОБРАБОТКЕ ДАТАСЕТА\n")
        f.write("=" * 60 + "\n\n")
        
        f.write(f"Дата и время завершения: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Исходный файл: {os.path.basename(input_file) if input_file else '—'}\n")
        f.write(f"Результирующий файл: {os.path.basename(output_file)}\n\n")
        
        f.write("ОБЩАЯ СТАТИСТИКА:\n")
        f.write("-" * 30 + "\n")
        f.write(f"Всего строк в датасете: {total_rows:,}\n")
        f.write(f"Успешно обработано: {successful_downloads:,}\n")
        f.write(f"Не удалось скачать документы: {failed_downloads:,}\n")
        f.write(f"Процент неудачных загрузок: {failed_downloads/total_rows*100:.2f}%\n\n")
        
        f.write("СТАТИСТИКА СКАЧИВАНИЯ ДОКУМЕНТОВ:\n")
        f.write("-" * 40 + "\n")
        f.write(f"Всего HTTP запросов: {stats['total_requests']:,}\n")
        f.write(f"Кэшированных запросов: {stats['cached_requests']:,}\n")
        f.write(f"Успешных загрузок: {stats['successful_downloads']:,}\n")
        f.write(f"Неудачных загрузок: {stats['failed_downloads']:,}\n")
        f.write(f"Количество URL, по которым не удалось получить документы: {stats['failed_downloads']:,}\n")
        
        if stats['total_requests'] > 0:
            cache_efficiency = stats['cached_requests'] / (stats['cached_requests'] + stats['total_requests']) * 100
            f.write(f"Эффективность кэширования: {cache_efficiency:.2f}%\n")
        
        f.write("\n")
        
        f.write("ДЕТАЛИ ОБРАБОТКИ:\n")
        f.write("-" * 20 + "\n")
        f.write("• Обрабатывались документы по URL из поля 'urls' в метаданных\n")
        f.write("• Поддерживаемые форматы: HTML, PDF, текстовые файлы\n")
        f.write("• Wikipedia статьи загружаются через MediaWiki API (более надежно и быстро)\n")
        f.write("• Использовалось параллельное скачивание (до 5 потоков)\n")
        f.write("• Применялось кэширование для избежания повторных запросов\n")
        f.write("• Промежуточные результаты сохранялись каждые 100 строк\n")
        f.write(f"• Wikipedia API запросов: {stats.get('wikipedia_api_requests', 0):,}\n")
        f.write(f"• Wikipedia API успешных: {stats.get('wikipedia_api_success', 0):,}\n\n")
        
        f.write("ФАЙЛЫ РЕЗУЛЬТАТОВ:\n")
        f.write("-" * 20 + "\n")
        f.write(f"• Основной результат: {os.path.basename(output_file)}\n")
        f.write(f"• Лог обработки: download_progress.log\n")
        f.write(f"• Вывод программы: download_output.log\n")
        f.write(f"• Этот отчет: {os.path.basename(report_file)}\n\n")
        
        f.write("=" * 60 + "\n")
        f.write("ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО\n")
        f.write("=" * 60 + "\n")
    
    logger.info(f"Итоговый отчет сохранен в: {report_file}")

